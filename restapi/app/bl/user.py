#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import division

from flask import g
from openpyxl import load_workbook
from app.models import User

import os
import hashlib
import re
import unicodedata

dir_path = os.path.dirname(os.path.realpath(__file__))

def read_data():
    arr = []
    try:
        wb = load_workbook(filename = dir_path + '/staff.xlsx')
        sheet_ranges = wb['staff']
        for row in sheet_ranges.iter_rows('D2:E125'):
            name = None
            email = None
            for cell in row:    
                if cell.column == 'D':
                    name = cell.value.strip()
                if cell.column == 'E':
                    email = cell.value.strip()

            password = hashlib.md5('123456').hexdigest()
            user = User(
                name=name,
                email=email,
                password=password
            )
            arr.append(user)
    except Exception as ex:
        print(str(ex))

    return arr


def no_accent_vietnamese(s):
    s = s.encode('utf-8').decode('utf-8') 
    s = re.sub(u'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', s)
    s = re.sub(u'[ÀÁẠẢÃĂẰẮẶẲẴÂẦẤẬẨẪ]', 'A', s)
    s = re.sub(u'èéẹẻẽêềếệểễ', 'e', s)
    s = re.sub(u'ÈÉẸẺẼÊỀẾỆỂỄ', 'E', s)
    s = re.sub(u'òóọỏõôồốộổỗơờớợởỡ', 'o', s)
    s = re.sub(u'ÒÓỌỎÕÔỒỐỘỔỖƠỜỚỢỞỠ', 'O', s)
    s = re.sub(u'ìíịỉĩ', 'i', s)
    s = re.sub(u'ÌÍỊỈĨ', 'I', s)
    s = re.sub(u'ùúụủũưừứựửữ', 'u', s)
    s = re.sub(u'ƯỪỨỰỬỮÙÚỤỦŨ', 'U', s)
    s = re.sub(u'ỳýỵỷỹ', 'y', s)
    s = re.sub(u'ỲÝỴỶỸ', 'Y', s)
    s = re.sub(u'Đ', 'D', s)
    s = re.sub(u'đ', 'd', s)

    return ''.join(c for c in unicodedata.normalize('NFD', s)
                  if unicodedata.category(c) != 'Mn')